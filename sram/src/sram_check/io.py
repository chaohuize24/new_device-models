from __future__ import annotations

import csv
import json
import math
from pathlib import Path
import re
from typing import Any

from .core import CrossSectionModel, DifferentialSpectrum


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def resolve_path(base: Path, text: str) -> Path:
    path = Path(text)
    return path if path.is_absolute() else (base / path).resolve()


def _numeric_rows_after(lines: list[str], marker: str) -> list[list[float]]:
    starts = [i for i, line in enumerate(lines) if marker in line]
    if not starts:
        raise ValueError(f"SPENVIS marker not found: {marker}")
    rows: list[list[float]] = []
    for line in lines[starts[0] + 1 :]:
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith(("'", "*")):
            if rows:
                break
            continue
        try:
            rows.append([float(item.strip()) for item in stripped.split(",")])
        except ValueError:
            if rows:
                break
    if len(rows) < 2:
        raise ValueError("SPENVIS spectrum has fewer than two numeric rows")
    return rows


def _load_spenvis_let_text(path: Path, spec: dict[str, Any]) -> DifferentialSpectrum:
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    mission_days = spec.get("mission_duration_days")
    if mission_days is None:
        match = next((re.search(r"'MIS_DUR'.*?([0-9.E+-]+)\s*,?'days'", line) for line in lines if "MIS_DUR" in line), None)
        if match is None:
            raise ValueError("mission duration not found; set mission_duration_days in config")
        mission_days = float(match.group(1))
    mission_s = float(mission_days) * 86400.0
    rows = _numeric_rows_after(
        lines,
        "'DFlux','(m!u-2!n sr!u-1!n) (MeV cm!u2!n g!u-1!n)!u-1!n'",
    )
    solid_angle = float(spec.get("solid_angle_sr", 4.0 * math.pi))
    let_mg = tuple(row[0] / 1000.0 for row in rows)
    integral_reference = rows[0][1] * 1.0e-4 * solid_angle / mission_s
    # dFluence/d(LET in g^-1) -> dFlux/d(LET in mg^-1): 1e-4*1000/mission_s.
    dflux = tuple(row[2] * 0.1 * solid_angle / mission_s for row in rows)
    return DifferentialSpectrum(
        x=let_mg,
        differential_flux=dflux,
        variable="let_mev_cm2_mg",
        x_unit="MeV cm^2/mg",
        flux_unit="cm^-2 s^-1 (MeV cm^2/mg)^-1",
        source=str(path),
        normalization=f"mission-average omnidirectional flux using {solid_angle:g} sr",
        integral_flux_reference=integral_reference,
    )


def _load_spenvis_proton_text(path: Path, spec: dict[str, Any]) -> DifferentialSpectrum:
    """Read the shielded proton-energy block embedded in SPENVIS long-term LET output."""
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    mission_days = spec.get("mission_duration_days")
    if mission_days is None:
        match = next(
            (
                re.search(r"'MIS_DUR'.*?([0-9.E+-]+)\s*,?'days'", line)
                for line in lines
                if "MIS_DUR" in line
            ),
            None,
        )
        if match is None:
            raise ValueError("mission duration not found; set mission_duration_days in config")
        mission_days = float(match.group(1))
    mission_s = float(mission_days) * 86400.0
    rows = _numeric_rows_after(
        lines,
        "'DFlux','m!u-2!n sr!u-1!n (MeV/n)!u-1!n',1,'Differential Fluence'",
    )
    solid_angle = float(spec.get("solid_angle_sr", 4.0 * math.pi))
    scale = 1.0e-4 * solid_angle / mission_s
    return DifferentialSpectrum(
        x=tuple(row[0] for row in rows),
        differential_flux=tuple(row[2] * scale for row in rows),
        variable="proton_energy_mev",
        x_unit="MeV",
        flux_unit="cm^-2 s^-1 MeV^-1",
        source=f"{path}#spacecraft-shielded-proton-spectrum",
        normalization=f"mission-average omnidirectional flux using {solid_angle:g} sr",
        integral_flux_reference=rows[0][1] * scale,
    )


def _load_normalized_xlsx(path: Path, spec: dict[str, Any]) -> DifferentialSpectrum:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("normalized_xlsx input requires openpyxl") from exc
    sheet_name = spec.get("sheet", "LET Spectrum")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise ValueError(f"empty sheet {sheet_name} in {path}")
    header = {str(value): index for index, value in enumerate(rows[0]) if value is not None}
    x_col = spec.get("x_column", "let_mev_cm2_mg")
    flux_col = spec.get(
        "flux_column", "differential_flux_cm2_s_per_mev_cm2_mg"
    )
    integral_col = spec.get("integral_column")
    for column in (x_col, flux_col):
        if column not in header:
            raise ValueError(f"missing column {column!r} in {sheet_name}")
    data = [row for row in rows[1:] if row[header[x_col]] is not None]
    integral = None
    if integral_col and integral_col in header and data and data[0][header[integral_col]] is not None:
        integral = float(data[0][header[integral_col]])
    return DifferentialSpectrum(
        x=tuple(float(row[header[x_col]]) for row in data),
        differential_flux=tuple(float(row[header[flux_col]]) for row in data),
        variable=spec.get("variable", "let_mev_cm2_mg"),
        x_unit=spec.get("x_unit", "MeV cm^2/mg"),
        flux_unit=spec.get("flux_unit", "cm^-2 s^-1 (MeV cm^2/mg)^-1"),
        source=f"{path}#{sheet_name}",
        normalization=spec.get("normalization", "already omnidirectional; no extra 4pi"),
        integral_flux_reference=integral,
    )


def _load_spectrum_csv(path: Path, spec: dict[str, Any]) -> DifferentialSpectrum:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    x_col = spec["x_column"]
    flux_col = spec["flux_column"]
    return DifferentialSpectrum(
        x=tuple(float(row[x_col]) for row in rows),
        differential_flux=tuple(float(row[flux_col]) for row in rows),
        variable=spec["variable"],
        x_unit=spec["x_unit"],
        flux_unit=spec["flux_unit"],
        source=str(path),
        normalization=spec["normalization"],
        integral_flux_reference=(
            None if spec.get("integral_flux_reference") is None else float(spec["integral_flux_reference"])
        ),
    )


def load_spectrum(spec: dict[str, Any], base: Path) -> DifferentialSpectrum:
    path = resolve_path(base, spec["path"])
    kind = spec["type"]
    if kind == "spenvis_let_text":
        return _load_spenvis_let_text(path, spec)
    if kind == "spenvis_proton_text":
        return _load_spenvis_proton_text(path, spec)
    if kind == "normalized_xlsx":
        return _load_normalized_xlsx(path, spec)
    if kind == "csv":
        return _load_spectrum_csv(path, spec)
    raise ValueError(f"unsupported spectrum type: {kind}")


def load_cross_section(spec: dict[str, Any], base: Path) -> CrossSectionModel:
    kind = spec["type"]
    common = {
        "variable": spec["variable"],
        "normalization": spec["normalization"],
        "source": spec.get("source", spec.get("path", "inline config")),
    }
    if kind == "weibull":
        return CrossSectionModel(
            kind="weibull",
            sigma_sat=float(spec["sigma_sat_cm2"]),
            threshold=float(spec["threshold"]),
            width=float(spec["width"]),
            shape=float(spec["shape"]),
            **common,
        )
    if kind != "table":
        raise ValueError(f"unsupported response type: {kind}")
    path = resolve_path(base, spec["path"])
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) < 2:
        raise ValueError(f"cross-section table needs at least two rows: {path}")
    x_col = spec["x_column"]
    sigma_col = spec["sigma_column"]
    low_col = spec.get("sigma_low_column")
    high_col = spec.get("sigma_high_column")
    return CrossSectionModel(
        kind="table",
        x=tuple(float(row[x_col]) for row in rows),
        sigma=tuple(float(row[sigma_col]) for row in rows),
        sigma_low=(tuple(float(row[low_col]) for row in rows) if low_col else None),
        sigma_high=(tuple(float(row[high_col]) for row in rows) if high_col else None),
        below_range=spec.get("below_range", "zero"),
        above_range=spec.get("above_range", "hold"),
        source=str(path),
        variable=common["variable"],
        normalization=common["normalization"],
    )
