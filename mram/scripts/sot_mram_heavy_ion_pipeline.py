#!/usr/bin/env python3
"""Generate a SOT-MRAM read-error cross-section from circuit Qcrit.

The calculation is deliberately event-coupled:

  circuit/mismatch sample + particle charge + injection node/time -> read result

It never combines a baseline circuit BER with a radiation BER as independent
events.  In the layout/TCAD-free phase it produces an engineering envelope,
not a foundry-calibrated absolute cross-section.

Two stages are supported:
  1. ``characterize``: Monte-Carlo ngspice binary searches of Qcrit at GBL/SA.
  2. ``cross-section``: RPP heavy-ion curve from those Qcrit samples and
     integration over the target-orbit LET spectrum.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import subprocess
import tempfile
from collections import defaultdict
from pathlib import Path

import numpy as np


MEASURE_RE = re.compile(r"^\s*([a-z0-9_]+)\s*=\s*([-+0-9.eE]+)\s*$", re.MULTILINE)


def resolve_ngspice(explicit: str | None, base: Path) -> str:
    if explicit:
        path = Path(explicit)
        if path.is_file():
            return str(path.resolve())
        sibling = path.parent / "ngspice_con.exe"
        if sibling.is_file():
            return str(sibling.resolve())
        return explicit
    env = os.environ.get("NGSPICE")
    if env and Path(env).is_file():
        return env
    cfg_path = base / "configs/pipeline_defaults.json"
    if cfg_path.is_file():
        cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
        for key in ("ngspice_executable", "ngspice_fallback"):
            candidate = cfg.get(key)
            if not candidate:
                continue
            path = Path(str(candidate))
            if path.is_file():
                con = path.parent / "ngspice_con.exe"
                return str(con.resolve()) if con.is_file() else str(path.resolve())
    return "ngspice"
Q_E_C = 1.602176634e-19
SI_EH_EV = 3.6
VTH_NOMINAL_V = 0.40


def _trapz(y: np.ndarray, x: np.ndarray) -> float:
    return float(np.trapezoid(y, x))


def load_spenvis_let_spectrum(path: Path) -> tuple[np.ndarray, np.ndarray, str]:
    """Load an LET rate spectrum in the model's final units.

    The normalized project workbook stores mission-averaged omnidirectional
    rates directly in MeV cm2/mg and cm^-2 s^-1 per (MeV cm2/mg).
    """
    if path.suffix.lower() != ".xlsx":
        raise ValueError("ROM heavy-ion integration requires the normalized SPENVIS workbook")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("reading the normalized SPENVIS workbook requires openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["LET Spectrum"]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if len(rows) < 2:
        raise ValueError(f"LET Spectrum is empty in {path}")
    header = {str(value): index for index, value in enumerate(rows[0]) if value is not None}
    required = {
        "let_mev_cm2_mg",
        "differential_flux_cm2_s_per_mev_cm2_mg",
    }
    missing = required - set(header)
    if missing:
        raise ValueError(f"missing columns in LET Spectrum: {sorted(missing)}")
    data = [row for row in rows[1:] if row[header["let_mev_cm2_mg"]] is not None]
    let_mg = np.asarray([float(row[header["let_mev_cm2_mg"]]) for row in data])
    dflux = np.asarray(
        [float(row[header["differential_flux_cm2_s_per_mev_cm2_mg"]]) for row in data]
    )
    normalization = "mission-average omnidirectional rate from LET Spectrum (already 4pi)"

    if let_mg.size < 2 or not np.all(np.isfinite(let_mg)) or not np.all(np.isfinite(dflux)):
        raise ValueError(f"invalid LET spectrum in {path}")
    if np.any(let_mg <= 0.0) or np.any(dflux < 0.0):
        raise ValueError(f"LET and differential flux must be non-negative in {path}")
    order = np.argsort(let_mg)
    let_mg = let_mg[order]
    dflux = dflux[order]
    if np.any(np.diff(let_mg) <= 0.0):
        raise ValueError(f"LET grid contains duplicate/non-increasing values in {path}")
    return let_mg, dflux, normalization


def set_param(source: str, name: str, value: str | float) -> str:
    # Parameters in the ROM netlist are sometimes grouped on one .param line.
    pattern = re.compile(rf"(\b{re.escape(name)}=)([^\s]+)", re.IGNORECASE)
    updated, count = pattern.subn(rf"\g<1>{value}", source, count=1)
    if count != 1:
        raise RuntimeError(f"parameter {name} was not found exactly once")
    return updated


def strip_control(source: str) -> str:
    return re.sub(
        r"\n\.control\n.*?\n\.endc\n",
        "\n.control\nrun\nquit\n.endc\n",
        source,
        count=1,
        flags=re.DOTALL | re.IGNORECASE,
    )


def positive_lognormal(rng: np.random.Generator, mean: float, relative_sigma: float) -> float:
    if relative_sigma <= 0:
        return mean
    sigma_ln = math.sqrt(math.log1p(relative_sigma**2))
    mu_ln = math.log(mean) - 0.5 * sigma_ln**2
    return float(rng.lognormal(mu_ln, sigma_ln))


def add_device_and_tid_envelope(
    source: str,
    *,
    dvth_n_mv: float,
    dvth_p_mv: float,
    mobility_scale: float,
    leakage_multiplier: float,
) -> str:
    block = (
        f"\n.param JOINT_DVTH_N_MV={dvth_n_mv:.9g}\n"
        f".param JOINT_DVTH_P_MV={dvth_p_mv:.9g}\n"
        f".param JOINT_MU_SCALE={mobility_scale:.9g}\n"
        f".param JOINT_LEAK_MULT={leakage_multiplier:.9g}\n"
    )
    source = source.replace(".param LCH=32n", block + ".param LCH=32n", 1)
    transformed: list[str] = []
    for line in source.splitlines():
        if line.startswith("M") and (" nmos " in line or " pmos " in line):
            shift = "JOINT_DVTH_N_MV" if " nmos " in line else "JOINT_DVTH_P_MV"
            if "DELVTO={" in line:
                line = re.sub(
                    r"DELVTO=\{([^}]*)\}",
                    rf"DELVTO={{(\1)+{shift}*1m}}",
                    line,
                    count=1,
                )
            else:
                line += f" DELVTO={{{shift}*1m}}"
            if "MULU0=" not in line:
                line += " MULU0={JOINT_MU_SCALE}"
        transformed.append(line)
    source = "\n".join(transformed) + "\n"
    leakage = """
VJOINT_LEAK_UNIT joint_leak_unit 0 {VDDVAL}
MJOINT_LEAK_UNIT joint_leak_unit 0 0 0 nmos W={WN} L={LCH} AD=0 AS=0 PD=0 PS=0 DELVTO={JOINT_DVTH_N_MV*1m} MULU0={JOINT_MU_SCALE}
BJOINT_EXTRA_D0 lbl_d0 0 I={NOFF*max((JOINT_LEAK_MULT-1)*(-i(VJOINT_LEAK_UNIT)),0)}
BJOINT_EXTRA_R0 lbl_r0 0 I={NOFF*max((JOINT_LEAK_MULT-1)*(-i(VJOINT_LEAK_UNIT)),0)}
BJOINT_EXTRA_D1 lbl_d1 0 I={NOFF*max((JOINT_LEAK_MULT-1)*(-i(VJOINT_LEAK_UNIT)),0)}
BJOINT_EXTRA_R1 lbl_r1 0 I={NOFF*max((JOINT_LEAK_MULT-1)*(-i(VJOINT_LEAK_UNIT)),0)}
"""
    return source.replace("VDD vdd 0 {VDDVAL}\n", "VDD vdd 0 {VDDVAL}\n" + leakage, 1)


def run_read1(
    template: str,
    circuit_dir: Path,
    ngspice: str,
    *,
    node: str,
    q_fc: float,
    injection_time_ns: float,
    n_present: int,
    par_scale: float,
    dvth_cell_mv: float,
    dvth_ref_mv: float,
    dvth_sa_mv: float,
    dvth_global_n_mv: float,
    dvth_global_p_mv: float,
    mobility_scale: float,
    leakage_multiplier: float,
) -> bool:
    source = set_param(template, "NPRESENT", n_present)
    source = set_param(source, "PAR_SCALE", f"{par_scale:.9g}")
    source = set_param(source, "DVTH_CELL_MV", f"{dvth_cell_mv:.9g}")
    source = set_param(source, "DVTH_REF_MV", f"{dvth_ref_mv:.9g}")
    source = set_param(source, "DVTH_SA_MV", f"{dvth_sa_mv:.9g}")
    source = set_param(source, "QGBL_FC", f"{q_fc:.9g}" if node == "gbl" else 0)
    source = set_param(source, "QSA_FC", f"{q_fc:.9g}" if node == "sa" else 0)
    source = set_param(source, "TINJ_GBL", f"{injection_time_ns:.9g}n")
    source = set_param(source, "TINJ_SA", f"{injection_time_ns:.9g}n")
    source = add_device_and_tid_envelope(
        source,
        dvth_n_mv=dvth_global_n_mv,
        dvth_p_mv=dvth_global_p_mv,
        mobility_scale=mobility_scale,
        leakage_multiplier=leakage_multiplier,
    )
    source = strip_control(source)

    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".cir", prefix="mram_joint_qcrit_", delete=False
    ) as handle:
        handle.write(source)
        netlist = Path(handle.name)
    try:
        completed = subprocess.run(
            [ngspice, "-b", str(netlist)],
            cwd=circuit_dir,
            text=True,
            capture_output=True,
            check=False,
        )
        log = completed.stdout + "\n" + completed.stderr
        if completed.returncode != 0:
            raise RuntimeError(log[-2500:])
        measures = {name.lower(): float(value) for name, value in MEASURE_RE.findall(log)}
        if "read1_ok" in measures:
            return bool(measures["read1_ok"] > 0.5)
        if "out1_sense" not in measures:
            raise RuntimeError(f"missing READ1_OK/OUT1_SENSE:\n{log[-2500:]}")
        return bool(measures["out1_sense"] > 0.5)
    finally:
        netlist.unlink(missing_ok=True)


def find_qcrit(
    template: str,
    circuit_dir: Path,
    ngspice: str,
    *,
    max_q_fc: float,
    tolerance_fc: float,
    **case,
) -> tuple[float, str]:
    if not run_read1(template, circuit_dir, ngspice, q_fc=0.0, **case):
        return 0.0, "baseline_failed"
    if run_read1(template, circuit_dir, ngspice, q_fc=max_q_fc, **case):
        return max_q_fc, "right_censored"
    lo, hi = 0.0, max_q_fc
    while hi - lo > tolerance_fc:
        mid = 0.5 * (lo + hi)
        if run_read1(template, circuit_dir, ngspice, q_fc=mid, **case):
            lo = mid
        else:
            hi = mid
    return hi, "resolved"


def write_tsv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"no rows for {path}")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]), delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def load_levels(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))["levels"]


def characterize(args: argparse.Namespace) -> None:
    base = Path(__file__).resolve().parents[1]
    ngspice = resolve_ngspice(args.ngspice, base)
    circuit_dir = base / "netlists"
    template = (circuit_dir / args.netlist).read_text(encoding="utf-8")
    if args.total_rows <= 0 or args.segment_rows <= 1 or args.array_columns <= 0:
        raise ValueError("array dimensions must be positive and segment_rows must exceed one")
    if args.total_rows % args.segment_rows:
        raise ValueError("total_rows must be an integer multiple of segment_rows")
    if args.array_columns % args.column_mux_ratio:
        raise ValueError("array_columns must be an integer multiple of column_mux_ratio")
    template = set_param(template, "TOTAL_ROWS", args.total_rows)
    template = set_param(template, "SEG_ROWS", args.segment_rows)
    template = set_param(template, "ARRAY_COLS", args.array_columns)
    template = set_param(template, "COLMUX", args.column_mux_ratio)
    levels = load_levels((base / args.envelopes).resolve())
    selected = args.level.split(",") if args.level != "all" else list(levels)
    unknown = set(selected) - set(levels)
    if unknown:
        raise ValueError(f"unknown envelope levels: {sorted(unknown)}")
    rng = np.random.default_rng(args.seed)
    times = {
        "gbl": [float(x) for x in args.gbl_times_ns.split(",")],
        "sa": [float(x) for x in args.sa_times_ns.split(",")],
    }
    rows: list[dict] = []
    for level_name in selected:
        level = levels[level_name]
        vth_sigma_mv = 1e3 * VTH_NOMINAL_V * level["vth_relative_sigma"]
        for sample in range(args.samples):
            static_sa_mv = rng.normal(0.0, level["sense_offset_sigma_mv"])
            mobility_process_scale = positive_lognormal(
                rng, 1.0, level["mobility_relative_sigma"]
            )
            common = {
                "n_present": int(rng.binomial(args.segment_rows - 1, args.present_probability)),
                "par_scale": positive_lognormal(rng, 1.0, level["capacitance_relative_sigma"]),
                "dvth_cell_mv": float(rng.normal(0.0, vth_sigma_mv)),
                "dvth_ref_mv": float(rng.normal(0.0, vth_sigma_mv)),
                "dvth_global_n_mv": level["tid_global_dvth_n_mv_at_14p32krad"],
                "dvth_global_p_mv": level["tid_global_dvth_p_mv_at_14p32krad"],
                "mobility_scale": mobility_process_scale
                * level["tid_mobility_scale_at_14p32krad"],
                "leakage_multiplier": level["tid_off_current_scale_at_14p32krad"],
            }
            for node, node_times in times.items():
                dynamic_noise_mv = rng.normal(0.0, level["dynamic_noise_sigma_mv"])
                common["dvth_sa_mv"] = float(static_sa_mv + dynamic_noise_mv)
                injection_time_ns = float(rng.choice(node_times))
                qcrit_fc, status = find_qcrit(
                    template,
                    circuit_dir,
                    ngspice,
                    node=node,
                    injection_time_ns=injection_time_ns,
                    max_q_fc=args.max_q_fc,
                    tolerance_fc=args.tolerance_fc,
                    **common,
                )
                rows.append(
                    {
                        "level": level_name,
                        "sample": sample,
                        "node": node,
                        "total_rows": args.total_rows,
                        "segment_rows": args.segment_rows,
                        "array_columns": args.array_columns,
                        "column_mux_ratio": args.column_mux_ratio,
                        "simultaneous_output_bits": args.array_columns // args.column_mux_ratio,
                        "injection_time_ns": injection_time_ns,
                        "qcrit_fc": qcrit_fc,
                        "qcrit_status": status,
                        "n_present": common["n_present"],
                        "par_scale": common["par_scale"],
                        "dvth_cell_mv": common["dvth_cell_mv"],
                        "dvth_ref_mv": common["dvth_ref_mv"],
                        "static_sa_offset_mv": static_sa_mv,
                        "dynamic_noise_mv": dynamic_noise_mv,
                        "equivalent_dvth_sa_mv": common["dvth_sa_mv"],
                        "vth_sigma_mv": vth_sigma_mv,
                        "mobility_process_scale": mobility_process_scale,
                        "combined_mobility_scale": common["mobility_scale"],
                        "tid_global_dvth_n_mv": common["dvth_global_n_mv"],
                        "tid_global_dvth_p_mv": common["dvth_global_p_mv"],
                        "tid_off_current_scale": common["leakage_multiplier"],
                    }
                )
    write_tsv((base / args.output).resolve(), rows)
    print(f"wrote {len(rows)} coupled Qcrit samples to {(base / args.output).resolve()}")


def charge_fc_from_let(let_mev_cm2_mg: np.ndarray, depth_um: float, efficiency: np.ndarray) -> np.ndarray:
    # Edep[MeV] = LET[MeV cm2/mg] * rho[mg/cm3] * depth[cm]
    edep_mev = let_mev_cm2_mg * 2330.0 * depth_um * 1e-4
    generated_fc = edep_mev * 1e6 / SI_EH_EV * Q_E_C * 1e15
    return generated_fc * efficiency


def sample_efficiency(level: dict, rng: np.random.Generator, count: int) -> np.ndarray:
    median = level["collection_efficiency_median"]
    cv = level["collection_efficiency_cv"]
    sigma_ln = math.sqrt(math.log1p(cv**2))
    return rng.lognormal(math.log(median), sigma_ln, count)


def beta_interval(successes: int, trials: int) -> tuple[float, float]:
    # Wilson interval; stable and dependency-free for zero counts.
    if trials <= 0:
        return 0.0, 1.0
    z = 1.959963984540054
    p = successes / trials
    den = 1.0 + z * z / trials
    center = (p + z * z / (2 * trials)) / den
    half = z * math.sqrt(p * (1 - p) / trials + z * z / (4 * trials * trials)) / den
    return max(0.0, center - half), min(1.0, center + half)


def heavy_ion_curve(qrows: list[dict[str, str]], levels: dict, rng: np.random.Generator, let_grid: np.ndarray) -> list[dict]:
    grouped: dict[tuple[str, str], list[float]] = defaultdict(list)
    for row in qrows:
        grouped[(row["level"], row["node"])].append(float(row["qcrit_fc"]))
    # Follow one Monte-Carlo device population across the full LET grid.
    # Re-drawing collection efficiency at every LET point can make a
    # physically monotonic cross-section appear noisy or non-monotonic.
    populations: dict[tuple[str, str], tuple[np.ndarray, np.ndarray]] = {}
    for key, values in grouped.items():
        qcrit = np.asarray(values, dtype=float)
        populations[key] = (qcrit, sample_efficiency(levels[key[0]], rng, qcrit.size))
    out: list[dict] = []
    for level_name in sorted({key[0] for key in grouped}):
        level = levels[level_name]
        for let_value in let_grid:
            total_sigma = total_low = total_high = 0.0
            total_trials = total_failures = 0
            for node in ("gbl", "sa"):
                qcrit, efficiency = populations[(level_name, node)]
                qcol = charge_fc_from_let(
                    np.full(qcrit.size, let_value), level["sensitive_depth_um"], efficiency
                )
                failures = int(np.count_nonzero(qcol >= qcrit))
                probability = failures / qcrit.size
                p_low, p_high = beta_interval(failures, qcrit.size)
                area_cm2 = level[f"{node}_sensitive_area_um2"] * 1e-8
                total_sigma += area_cm2 * probability
                total_low += area_cm2 * p_low
                total_high += area_cm2 * p_high
                total_trials += qcrit.size
                total_failures += failures
            out.append(
                {
                    "level": level_name,
                    "let_mev_cm2_mg": let_value,
                    "sigma_cm2_per_active_read_bit": total_sigma,
                    "sigma_ci95_low": total_low,
                    "sigma_ci95_high": total_high,
                    "failed_joint_samples": total_failures,
                    "joint_samples": total_trials,
                    "model": "ngspice_Qcrit_plus_RPP_charge_collection",
                }
            )
    return out


def integrate_spenvis_let(
    curve: list[dict],
    path: Path,
    read_window_ns: float,
    qrows: list[dict[str, str]],
    levels: dict,
    rng: np.random.Generator,
    bootstrap_samples: int,
) -> list[dict]:
    let_mg, dflux, spectrum_normalization = load_spenvis_let_spectrum(path)
    # Omnidirectional cumulative flux above any LET threshold.  This lets each
    # coupled qcrit sample contribute an exact step-response rate without
    # treating pointwise Wilson bounds as a global confidence interval.
    segment_integrals = 0.5 * (dflux[:-1] + dflux[1:]) * np.diff(let_mg)
    cumulative_above = np.zeros_like(let_mg)
    cumulative_above[:-1] = np.cumsum(segment_integrals[::-1])[::-1]

    rates: list[dict] = []
    for level_name in sorted({row["level"] for row in curve}):
        sub = [row for row in curve if row["level"] == level_name]
        x = np.asarray([row["let_mev_cm2_mg"] for row in sub])
        sigma = np.asarray([row["sigma_cm2_per_active_read_bit"] for row in sub])
        sigma_low = np.asarray([row["sigma_ci95_low"] for row in sub])
        sigma_high = np.asarray([row["sigma_ci95_high"] for row in sub])
        def integrate(values: np.ndarray) -> float:
            interp = np.interp(let_mg, x, values, left=values[0], right=values[-1])
            return _trapz(dflux * interp, let_mg)
        curve_rate = integrate(sigma)
        rate_low_pointwise = integrate(sigma_low)
        rate_high_pointwise = integrate(sigma_high)

        level = levels[level_name]
        sample_rows: dict[int, dict[str, float]] = defaultdict(dict)
        for row in qrows:
            if row["level"] == level_name:
                sample_rows[int(row["sample"])][row["node"]] = float(row["qcrit_fc"])
        contributions: list[float] = []
        for sample in sorted(sample_rows):
            total = 0.0
            for node in ("gbl", "sa"):
                efficiency = float(sample_efficiency(level, rng, 1)[0])
                charge_per_let_fc = float(
                    charge_fc_from_let(
                        np.asarray([1.0]), level["sensitive_depth_um"], np.asarray([efficiency])
                    )[0]
                )
                threshold = sample_rows[sample][node] / charge_per_let_fc
                flux_above = float(
                    np.interp(threshold, let_mg, cumulative_above, left=cumulative_above[0], right=0.0)
                )
                total += level[f"{node}_sensitive_area_um2"] * 1e-8 * flux_above
            contributions.append(total)
        contribution_array = np.asarray(contributions)
        rate = float(np.mean(contribution_array))
        boot = np.empty(bootstrap_samples)
        for index in range(bootstrap_samples):
            boot[index] = float(np.mean(rng.choice(contribution_array, contribution_array.size, replace=True)))
        rate_low, rate_high = np.quantile(boot, [0.025, 0.975])
        window_s = read_window_ns * 1e-9
        rates.append(
            {
                "level": level_name,
                "heavy_ion_set_rate_per_continuously_sensitive_bit_s": rate,
                "rate_ci95_bootstrap_low": rate_low,
                "rate_ci95_bootstrap_high": rate_high,
                "curve_grid_integral_rate_s": curve_rate,
                "rate_low_pointwise_wilson_diagnostic": rate_low_pointwise,
                "rate_high_pointwise_wilson_diagnostic": rate_high_pointwise,
                "read_window_ns": read_window_ns,
                "probability_per_active_bit_read": -math.expm1(-rate * window_s),
                "probability_ci95_bootstrap_low": -math.expm1(-rate_low * window_s),
                "probability_ci95_bootstrap_high": -math.expm1(-rate_high * window_s),
                "confidence_scope": "paired-sample bootstrap within each envelope; excludes envelope/geometry/model uncertainty",
                "coupled_samples": contribution_array.size,
                "bootstrap_resamples": bootstrap_samples,
                "spectrum": "environment/target_orbit/SPENVIS_环境与模型输入.xlsx",
                "spectrum_normalization": spectrum_normalization,
                "solid_angle_assumption": "4pi omnidirectional; not applied twice",
            }
        )
    return rates


def cross_section(args: argparse.Namespace) -> None:
    base = Path(__file__).resolve().parents[1]
    levels = load_levels((base / args.envelopes).resolve())
    qrows = read_tsv((base / args.qcrit_input).resolve())
    # Keep the heavy-ion curve and LET-rate bootstrap reproducible independently.
    seed_hi, seed_rate = np.random.SeedSequence(args.seed).spawn(2)
    rng_hi = np.random.default_rng(seed_hi)
    rng_rate = np.random.default_rng(seed_rate)
    let_grid = np.geomspace(args.let_min, args.let_max, args.let_points)
    hi_rows = heavy_ion_curve(qrows, levels, rng_hi, let_grid)
    write_tsv((base / args.heavy_ion_output).resolve(), hi_rows)
    if args.spenvis_let:
        rates = integrate_spenvis_let(
            hi_rows,
            (base / args.spenvis_let).resolve(),
            args.read_window_ns,
            qrows,
            levels,
            rng_rate,
            args.bootstrap_samples,
        )
        write_tsv((base / args.rate_output).resolve(), rates)
    print(f"wrote heavy-ion curve to {(base / args.heavy_ion_output).resolve()}")


def make_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    char = sub.add_parser("characterize")
    char.add_argument("--netlist", default="sot_mram_hierarchical_senseamp.cir")
    char.add_argument("--ngspice", default=None, help="path to ngspice or ngspice_con executable")
    char.add_argument("--envelopes", default="configs/statistical_envelopes.json")
    char.add_argument("--level", default="all", help="low,nominal,high or all")
    char.add_argument("--samples", type=int, default=64)
    char.add_argument("--seed", type=int, default=20260826)
    char.add_argument("--present-probability", type=float, default=0.5)
    char.add_argument("--total-rows", type=int, default=2048)
    char.add_argument("--segment-rows", type=int, default=256)
    char.add_argument("--array-columns", type=int, default=128)
    char.add_argument("--column-mux-ratio", type=int, default=4)
    char.add_argument("--gbl-times-ns", default="1.10,1.25,1.40,1.55,1.62")
    char.add_argument("--sa-times-ns", default="2.05,2.15,2.25,2.35,2.42")
    char.add_argument("--max-q-fc", type=float, default=200.0)
    char.add_argument("--tolerance-fc", type=float, default=0.5)
    char.add_argument("--output", default="results/mram_joint_qcrit_samples.tsv")
    char.set_defaults(func=characterize)

    cross = sub.add_parser("cross-section")
    cross.add_argument("--envelopes", default="configs/statistical_envelopes.json")
    cross.add_argument("--qcrit-input", default="results/mram_joint_qcrit_samples.tsv")
    cross.add_argument("--seed", type=int, default=20260826)
    cross.add_argument("--let-min", type=float, default=0.01)
    cross.add_argument("--let-max", type=float, default=120.0)
    cross.add_argument("--let-points", type=int, default=100)
    cross.add_argument("--heavy-ion-output", default="results/mram_heavy_ion_cross_section.tsv")
    cross.add_argument(
        "--spenvis-let",
        default="../environment/target_orbit/SPENVIS_环境与模型输入.xlsx",
        help=(
            "normalized SPENVIS workbook (LET Spectrum sheet); the default "
            "is the formal 5-year, 1 g/cm2 Al mission environment"
        ),
    )
    cross.add_argument("--rate-output", default="results/mram_spenvis_heavy_ion_rate.tsv")
    cross.add_argument("--read-window-ns", type=float, default=0.72)
    cross.add_argument("--bootstrap-samples", type=int, default=10000)
    cross.set_defaults(func=cross_section)
    return parser


def main() -> None:
    args = make_parser().parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
